import openpyxl
import json
from datetime import datetime
import time

def parse_last_name(full_name):
    if not full_name:
        return ""
    # Format is "LastName, FirstName "
    if ',' in full_name:
        return full_name.split(',')[0].strip()
    return full_name.strip()

def parse_team_name(name):
    if not name:
        return ""
    name = str(name).strip()
    if ',' in name:
        parts = [p.strip() for p in name.split(',')]
        if len(parts) >= 2:
            # Swap part after first comma to be first
            return f"{parts[1]} {parts[0]}"
    return name

def process_date(dt):
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d')
    if isinstance(dt, str):
        # Already string? Try parsing common formats if needed
        return dt.split(' ')[0] # Simple split for "YYYY-MM-DD HH:MM:SS"
    return str(dt)

def main():
    wb = openpyxl.load_workbook('/Users/harryscons/Documents/celestial-plasma/Relays.xlsx', data_only=True)
    if 'RelayData' in wb.sheetnames:
        sheet = wb['RelayData']
    else:
        sheet = wb.active
    
    cols = [cell.value for cell in sheet[1]]
    records = []
    
    # Base timestamp to avoid collisions
    base_ts = int(time.time() * 1000)
    
    for i, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not any(row_values):
            continue
            
        row = dict(zip(cols, row_values))
        
        # Flexibly find Athlete columns
        def get_val_flex(prefixes):
            for p in prefixes:
                for k in row.keys():
                    if k and str(k).lower().strip().replace(' ', '') == p.lower().replace(' ', ''):
                        return row.get(k)
            return None

        # Try common variants of "Athlete 1"
        p1 = str(get_val_flex(['Athlete01', 'Athlete 1', 'Atlhlete 01', 'Athllete 1', 'Αθλητής 1']) or "").strip()
        p2 = str(get_val_flex(['Athlete02', 'Athlete 2', 'Atlhlete 02', 'Athllete 2', 'Αθλητής 2']) or "").strip()
        p3 = str(get_val_flex(['Athlete03', 'Athlete 3', 'Atlhlete 03', 'Athllete 3', 'Αθλητής 3']) or "").strip()
        p4 = str(get_val_flex(['Athlete04', 'Athlete 4', 'Atlhlete 04', 'Athllete 4', 'Αθλητής 4']) or "").strip()
        
        participants = [p for p in [p1, p2, p3, p4] if p]
        
        # Notes: Concatenate Last Names
        last_names = [parse_last_name(p) for p in participants if p]
        notes = "/".join(last_names)
        
        record = {
            "id": base_ts + i,
            "event": str(row.get('Event') or ""),
            "gender": str(row.get('Gender') or ""),
            "ageGroup": str(row.get('Age Group') or ""),
            "athlete": parse_team_name(row.get('Team Name')),
            "isRelay": True,
            "relayParticipants": participants,
            "raceName": str(row.get('Race Name') or ""),
            "idr": str(row.get('IDR') or ""),
            "notes": notes,
            "mark": str(row.get('Mark') or ""),
            "wind": "", # Not in Excel?
            "date": process_date(row.get('Date')),
            "town": str(row.get('Town') or ""),
            "country": "Ελλάδα" # Defaulting if not specified, though Town is usually Greek
        }
        records.append(record)
        
    with open('/Users/harryscons/Documents/celestial-plasma/imported_relays.json', 'w', encoding='utf-8') as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
        
    print(f"Successfully processed {len(records)} records.")

if __name__ == "__main__":
    main()
