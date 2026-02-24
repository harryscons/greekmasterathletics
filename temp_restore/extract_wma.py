import openpyxl
import json
import time

def extract_sheet(sheet, gender_label):
    data = []
    # Row 0 is Measure, Row 1 is Event names
    rows = list(sheet.iter_rows(values_only=True))
    events = rows[1][1:] # List of event names from index 1 to 30
    
    for row in rows[2:]: # Start from data rows
        age = row[0]
        if age is None or not isinstance(age, (int, float)):
            continue
        
        age = int(age)
        if age < 35: continue # Usually masters start at 35, but tables might have 30
        
        # Determine Age Group (e.g., M35, W40)
        prefix = 'M' if gender_label == 'men' else 'W'
        age_group = f"{prefix}{ (age // 5) * 5 }"
        
        for i, factor in enumerate(row[1:]):
            event_name = events[i]
            if event_name is None: continue
            if factor is None: continue
            
            data.append({
                "id": int(time.time() * 1000) + len(data),
                "gender": gender_label,
                "ageGroup": age_group,
                "event": event_name,
                "age": age,
                "factor": float(factor)
            })
    return data

def main():
    wb = openpyxl.load_workbook('Appendix-B_2023.xlsx', data_only=True)
    
    all_data = []
    
    if 'Male' in wb.sheetnames:
        print("Extracting Male sheet...")
        all_data.extend(extract_sheet(wb['Male'], 'men'))
        
    if 'Female' in wb.sheetnames:
        print("Extracting Female sheet...")
        all_data.extend(extract_sheet(wb['Female'], 'women'))
        
    with open('wma_data_2023.json', 'w') as f:
        json.dump(all_data, f, indent=2)
        
    print(f"Extracted {len(all_data)} records to wma_data_2023.json")

if __name__ == "__main__":
    main()
